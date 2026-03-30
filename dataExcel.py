import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk
import os
import threading
import time
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# ==============================================================================
# 配置区域：分校分类与固定顺序
# ==============================================================================
FIXED_ORDER_CITIES = [
    "新乡", "沧州", "松原", "临汾", "汉中", "朝阳", "徐州",
    "海淀", "怀柔", "顺义", "房山", "大兴", "昌平", "平谷", "通州", "延庆", "密云",
    "包头", "通辽", "赤峰",
    "天津", "武清", "静海", "滨海"
]

DEFAULT_BRANCH_MAP = {
    "新乡": "河南分校", "沧州": "河北分校", "松原": "吉林分校", "临汾": "山西分校",
    "汉中": "陕西分校", "朝阳": "辽宁分校", "徐州": "江苏分校",
    "海淀": "北京分校", "怀柔": "北京分校", "顺义": "北京分校", "房山": "北京分校",
    "大兴": "北京分校", "昌平": "北京分校", "平谷": "北京分校", "通州": "北京分校",
    "延庆": "北京分校", "密云": "北京分校",
    "包头": "内蒙古分校", "通辽": "内蒙古分校", "赤峰": "内蒙古分校",
    "天津": "天津分校", "武清": "天津分校", "静海": "天津分校", "滨海": "天津分校"
}

BRANCH_GROUPS = {
    "A": ["山东分校", "广东分校", "河南分校", "河北分校", "湖北分校", "吉林分校", "山西分校", "陕西分校", "安徽分校", "辽宁分校", "云南分校"],
    "B": ["江苏分校", "湖南分校", "四川分校", "黑龙江分校", "广西分校", "新疆分校", "浙江分校", "江西分校",  "北京分校","内蒙古分校"],
    "C": ["贵州分校","福建分校","甘肃分校", "海南分校",  "宁夏分校", "青海分校", "厦门分校", "上海分校", "天津分校", "西藏分校", "重庆分校"]
}

# ==============================================================================
# 核心处理类
# ==============================================================================
class UniversalProcessor:
    def __init__(self, log_func):
        self.log_callback = log_func
        self.use_special_city_logic = False

    def log(self, message):
        """带有 UI 强制刷新的日志输出"""
        self.log_callback(message)

    def excel_lookup_find(self, text, keywords, results, default_value):
        """模拟 Excel 的 LOOKUP FIND 逻辑，用于识别分校"""
        if pd.isna(text):
            return default_value
        
        text_string = str(text)
        # 逆序查找，确保长关键词优先匹配
        for keyword, result in zip(reversed(keywords), reversed(results)):
            if keyword in text_string:
                return result
        return default_value

    def check_keyword_flag(self, text, keywords):
        """检查文本是否包含关键词列表中的任何一个，返回 1 或 空"""
        if pd.isna(text):
            return ""
        
        text_string = str(text)
        for keyword in keywords:
            if keyword in text_string:
                return 1
        return ""

    def extract_city_smart(self, dept_path):
        """从负责人所属部门路径中智能提取地市名称 (融合 4.0 核心逻辑回归版)"""
        if pd.isna(dept_path):
            return "未知"
        
        text_string = str(dept_path).replace('\r', '\n')
        lines = [line.strip() for line in text_string.split('\n') if line.strip()]
        if not lines:
            return "其他"
            
        target_line = ""
        for line in lines:
            if "地市分校" in line:
                target_line = line
                break
        if not target_line:
            for line in lines:
                if "分校" in line:
                    target_line = line
                    break
        if not target_line:
            target_line = lines[0]
            
        clean_text = target_line.replace('"', '').replace("'", "").replace("华图教育", "").strip()
        parts = [part.strip() for part in clean_text.split('/') if part.strip()]
        if not parts:
            return "其他"
            
        final_city_name = ""
        # 1. 直接匹配包含"地市分校"的节点
        for part in parts:
            if "地市分校" in part and part != "地市分校":
                final_city_name = part.replace("地市分校", "")
                break
                
        # 2. ★ 找回 4.0 丢失的逻辑：处理 "XX分校/各校区/某某市" 的多层级嵌套
        if not final_city_name:
            for i, part in enumerate(parts):
                if "分校" in part and len(part) > 2:
                    if i + 1 < len(parts):
                        target = parts[i+1]
                        # 智能跳过无意义的中间层级
                        if target in ["地市分校", "各校区"]:
                            if i + 2 < len(parts): 
                                target = parts[i+2]
                            else: 
                                continue
                        final_city_name = target
                        break
                        
        # 3. 兜底：取最后一级
        if not final_city_name:
            final_city_name = parts[-1]
            
        # 4. 剥离冗余后缀 (保留 6.0 的 while 循环优势，彻底清除叠加后缀)
        suffixes = ["区属学习中心", "高校学习中心", "学习中心", "办事处", "地市分校", "分校", "分部", "校区", "运营中心", "基地", "工作站", "旗舰店"]
        changed = True
        while changed:
            changed = False
            for suffix in suffixes:
                if final_city_name.endswith(suffix):
                    final_city_name = final_city_name[:-len(suffix)]
                    changed = True
                    
        if not final_city_name:
            return "其他"
            
        return final_city_name

    def standardize_city_name(self, city_name):
        """标准化地市名称，去除市字和空白符 (找回 4.0 的特殊字符清理)"""
        if pd.isna(city_name):
            return "其他"
        
        standard_name = str(city_name).strip()
        
        # ★ 找回 4.0 丢失的逻辑：清理 Excel 导出的底层换行符编码
        standard_name = standard_name.replace("_x000D_", "").replace("_x000d_", "")
        
        # 清除所有不可见空白符
        standard_name = re.sub(r'[\s\r\n\t\u200b\ufeff\xa0]+', '', standard_name)
        
        # 去除末尾的"市"字（排除误伤如"沙市"）
        if len(standard_name) > 2 and standard_name.endswith("市"):
            standard_name = standard_name[:-1]
            
        if not standard_name:
            return "其他"
            
        return standard_name

    def process_step1(self, source_df, date_keyword, custom_total_kws_str):
        """步骤1：全量数据清洗与业务标签计算 (严格回归 4.0 逻辑)"""
        self.log("【数据清洗】正在提取地市与计算业务标签...")
        
        # 统一表头格式
        source_df.columns = [str(column).replace('\ufeff', '').strip() for column in source_df.columns]
        
        # 识别分校
        branch_keywords = ["通辽分校", "陕西分校", "湖北分校", "辽宁分校", "河北分校", "甘肃分校", "厦门分校", "福建分校", "山东分校", "北京分校", "安徽分校", "黑龙江分校", "吉林分校", "江苏分校", "重庆分校", "广东分校", "天津分校", "河南分校", "云南分校", "江西分校", "湖南分校", "贵州分校", "广西分校", "山西分校", "宁夏分校", "内蒙古分校", "浙江分校", "新疆分校", "青海分校", "南疆分校", "四川分校", "上海分校", "海南分校", "西藏分校", "赤峰"]
        branch_results = ["内蒙古分校", "陕西分校", "湖北分校", "辽宁分校", "河北分校", "甘肃分校", "厦门分校", "福建分校", "山东分校", "北京分校", "安徽分校", "黑龙江分校", "吉林分校", "江苏分校", "重庆分校", "广东分校", "天津分校", "河南分校", "云南分校", "江西分校", "湖南分校", "贵州分校", "广西分校", "山西分校", "宁夏分校", "内蒙古分校", "浙江分校", "新疆分校", "青海分校", "新疆分校", "四川分校", "上海分校", "海南分校", "西藏分校", "内蒙古分校"]
        
        source_df['分校'] = source_df['负责人所属部门'].apply(lambda x: self.excel_lookup_find(x, branch_keywords, branch_results, "总部"))
        source_df['地市'] = source_df['负责人所属部门'].apply(self.extract_city_smart)
        source_df['地市'] = source_df['地市'].apply(self.standardize_city_name)

        # 类别判断
        def determine_category(row):
            dept_raw = str(row['负责人所属部门'])
            if "地市分校" in dept_raw:
                return "地市"
            if str(row['分校']) in ["北京分校", "天津分校"] and "各校区" in dept_raw:
                return "地市"
            return "其它"
            
        source_df['所属类别'] = source_df.apply(determine_category, axis=1)

        # 核心业务标签打标
        source_df['公职'] = source_df['备注名'].apply(lambda x: self.check_keyword_flag(x, ["26公职","27公职","28公职","29公职"]))
        kw_shiye = ["26事业","26三支","26社区","26辅警","26书记员","26国企","30事业","30三支","30社区","30辅警","30书记员","30国企", "27事业","27三支","27社区","27辅警","27书记员","27国企","28事业","28三支","28社区","28辅警","28书记员","28国企", "29事业","29三支","29社区","29辅警","29书记员","29国企"]
        source_df['事业辅助列'] = source_df['备注名'].apply(lambda x: self.check_keyword_flag(x, kw_shiye))
        source_df['教师'] = source_df['备注名'].apply(lambda x: self.check_keyword_flag(x, ["26教师","26特岗","26教资","30教师","30特岗","30教资","27教师","27特岗","27教资","28教师","28特岗","28教资","29教师","29特岗","29教资"]))
        source_df['文职'] = source_df['备注名'].apply(lambda x: self.check_keyword_flag(x, ["30文职","26文职","27文职","28文职","29文职"]))
        source_df['医疗'] = source_df['备注名'].apply(lambda x: self.check_keyword_flag(x, ["30医疗","26医疗","27医疗","28医疗","29医疗"]))
        source_df['银行'] = source_df['备注名'].apply(lambda x: self.check_keyword_flag(x, ["30银行","26银行","27银行","28银行","29银行"]))
        source_df['考研'] = source_df['备注名'].apply(lambda x: self.check_keyword_flag(x, ["30考研","26考研","27考研","28考研","29考研"]))
        source_df['学历'] = source_df['备注名'].apply(lambda x: self.check_keyword_flag(x, ["30学历","26学历","27学历","28学历","29学历"]))
        source_df['其他'] = ""
        
        # 标备总数判定
        if custom_total_kws_str.strip():
            custom_keywords = [k.strip() for k in custom_total_kws_str.replace('，', ',').split(',') if k.strip()]
            def check_custom_total(remark):
                if pd.isna(remark): return 0
                if any(kw in str(remark) for kw in custom_keywords):
                    return 1
                return 0
            source_df['标备总数'] = source_df['备注名'].apply(check_custom_total)
        else:
            def check_default_total(row):
                standard_columns = ['公职', '事业辅助列', '教师', '文职', '医疗', '银行', '考研', '学历']
                for col in standard_columns:
                    if row[col] == 1:
                        return 1
                return 0
            source_df['标备总数'] = source_df.apply(check_default_total, axis=1)

        # 日期判定
        date_keywords_list = [k.strip() for k in date_keyword.replace('，', ',').split(',') if k.strip()]
        if not date_keywords_list:
            date_keywords_list = ["FAIL_SAFE"]
            
        def check_is_new_friend(create_time):
            time_str = str(create_time)
            if any(k in time_str for k in date_keywords_list):
                return "是"
            return "否"
            
        source_df['！！！是否新增'] = source_df['创建时间'].apply(check_is_new_friend)
        
        def format_date_str(val):
            try:
                dt_obj = pd.to_datetime(val)
                return f"{dt_obj.month}月{dt_obj.day}日"
            except:
                return str(val)
        source_df['日期'] = source_df['好友添加时间'].apply(format_date_str)

        # ★ 恢复 4.0 完整的 8词 渠道映射字典
        channel_keys = ["网络","社会","高校","线上","线上平台","线上活动","现场","线下活动"]
        channel_results = ["线上平台","线下活动","高校","线上平台","线上平台","线上平台","考试现场","线下活动"]
        source_df['渠道'] = source_df['渠道活码分组'].apply(lambda x: self.excel_lookup_find(x, channel_keys, channel_results, "其他"))
        
        online_keys = ["网站","公众号","小红书","视频号","抖音","文章页","网页","专题","附件","小程序"]
        online_results = ["网站","公众号","小红书","视频号","抖音","网站","网站","网站","网站","其他"]
        source_df['线上渠道'] = source_df['渠道活码分组'].apply(lambda x: self.excel_lookup_find(x, online_keys, online_results, "其他"))

        def check_shiye_sub(row):
            if row['公职'] != 1 and row['事业辅助列'] == 1:
                return 1
            return 0
        source_df['事业2'] = source_df.apply(check_shiye_sub, axis=1)
        
        def check_client_reply(row):
            has_reply_time = pd.notna(row['客户上次回复时间']) and str(row['客户上次回复时间']).strip() != ""
            if has_reply_time and row['标备总数'] == 1:
                return "是"
            return "否"
        source_df['客户回话'] = source_df.apply(check_client_reply, axis=1)

        # 强制挂载质检原生 5 字段
        qc_original_fields = ['好友添加来源', '添加渠道码', '员工发送消息数', '客户回复消息数', '是否同意会话存档']
        for field_name in qc_original_fields:
            if field_name not in source_df.columns:
                if '数' in field_name:
                    source_df[field_name] = 0
                else:
                    source_df[field_name] = "未知"

        # ★ 恢复 4.0 严格去重基准：强制提取第 26 列 (索引为25) 作为 ID
        id_column_name = None
        if len(source_df.columns) > 25:
            id_column_name = source_df.columns[25]
            
        if 'ExternalUserId' not in source_df.columns:
            if id_column_name:
                source_df.rename(columns={id_column_name: 'ExternalUserId'}, inplace=True)
            elif len(source_df.columns) > 1:
                # 兜底逻辑：如果没有26列，取第2列
                source_df.rename(columns={source_df.columns[1]: 'ExternalUserId'}, inplace=True)

        final_extract_cols = [
            '备注名', 'ExternalUserId', '分校', '地市', '所属类别', 
            '公职', '事业2', '教师', '文职', '医疗', '银行', '考研', '学历', '其他', 
            '标备总数', '！！！是否新增', '渠道', '线上渠道', '客户回话', '日期', 
            '添加好友状态', '好友添加来源', '添加渠道码', '员工发送消息数', '客户回复消息数', '是否同意会话存档'
        ]
        
        actual_cols_present = [c for c in final_extract_cols if c in source_df.columns]
        df_cleaned_all = source_df[actual_cols_present].copy()
        
        # 强制转换数值列以防计算报错
        for numeric_col in ['公职', '事业2', '标备总数', '员工发送消息数', '客户回复消息数']:
            if numeric_col in df_cleaned_all.columns:
                df_cleaned_all[numeric_col] = pd.to_numeric(df_cleaned_all[numeric_col], errors='coerce').fillna(0)

        self.log("正在执行全量数据去重逻辑与留存筛选...")
        df_dedup_province = df_cleaned_all.drop_duplicates(subset=['ExternalUserId', '分校'], keep='first')
        df_dedup_city = df_cleaned_all.drop_duplicates(subset=['ExternalUserId', '分校', '地市'], keep='first')
        
        df_retention_only = pd.DataFrame()
        if '添加好友状态' in df_cleaned_all.columns:
            df_retention_only = df_cleaned_all[df_cleaned_all['添加好友状态'].astype(str).str.contains("已添加", na=False)].copy()

        return df_cleaned_all, df_dedup_province, df_dedup_city, df_retention_only

    # --- 统计辅助函数 (保持 4.0 逻辑清晰度) ---
    def calculate_summary_stats(self, raw_data_df, dedup_data_df):
        """计算基础统计指标：新增、去重、标备、回话"""
        new_added_subset = dedup_data_df[dedup_data_df['！！！是否新增'].astype(str).str.contains("是", na=False)]
        
        raw_count_val = len(raw_data_df)
        dedup_count_val = len(dedup_data_df)
        new_added_count_val = len(new_added_subset)
        
        gz_sum_val = new_added_subset['公职'].sum()
        sy_sum_val = new_added_subset['事业2'].sum()
        total_tb_sum_val = new_added_subset['标备总数'].sum()
        
        reply_count_val = len(new_added_subset[new_added_subset['客户回话'].astype(str).str.contains("是", na=False)])
        
        results = {
            "raw_count": raw_count_val,
            "dedup_count": dedup_count_val,
            "new_added_count": new_added_count_val,
            "gz_sum": gz_sum_val,
            "sy_sum": sy_sum_val,
            "total_tb_sum": total_tb_sum_val,
            "reply_count": reply_count_val
        }
        return results

    def format_output_row_list(self, stats_dict, is_province_mode=False):
        """将统计字典转换为报表行，强制整数百分比"""
        def safe_division(a, b):
            return a / b if b != 0 else 0.0
            
        def to_percentage_integer(value):
            return f"{int(round(value * 100))}%"

        repeat_rate = safe_division((stats_dict["dedup_count"] - stats_dict["new_added_count"]), stats_dict["dedup_count"])
        other_tb_count = stats_dict["total_tb_sum"] - stats_dict["gz_sum"] - stats_dict["sy_sum"]
        tb_rate = safe_division(stats_dict["total_tb_sum"], stats_dict["new_added_count"])
        reply_rate = safe_division(stats_dict["reply_count"], stats_dict["new_added_count"])
        
        return [
            int(stats_dict["raw_count"]),
            int(stats_dict["dedup_count"]),
            int(stats_dict["new_added_count"]),
            to_percentage_integer(repeat_rate),
            int(stats_dict["gz_sum"]),
            int(stats_dict["sy_sum"]),
            int(other_tb_count),
            int(stats_dict["total_tb_sum"]),
            to_percentage_integer(tb_rate),
            int(stats_dict["reply_count"]),
            to_percentage_integer(reply_rate)
        ]

    def calc_stats_retention_data(self, retention_df_subset):
        """计算留存版专用指标"""
        if len(retention_df_subset) == 0:
            return {"gz": 0, "sy": 0, "tb": 0, "ret": 0}
        
        gz_sum_val = retention_df_subset['公职'].sum()
        sy_sum_val = retention_df_subset['事业2'].sum()
        tb_sum_val = retention_df_subset['标备总数'].sum()
        
        return {
            "gz": gz_sum_val,
            "sy": sy_sum_val,
            "tb": tb_sum_val,
            "ret": len(retention_df_subset)
        }

    def format_retention_row_list(self, stats_dict):
        """留存版行数据转换"""
        def safe_division(a, b):
            return a / b if b != 0 else 0.0
            
        def to_percentage_integer(value):
            return f"{int(round(value * 100))}%"
            
        other_count = stats_dict["tb"] - stats_dict["gz"] - stats_dict["sy"]
        tb_rate = safe_division(stats_dict["tb"], stats_dict["ret"])
        
        return [
            int(stats_dict["ret"]),
            int(stats_dict["gz"]),
            int(stats_dict["sy"]),
            int(other_count),
            int(stats_dict["tb"]),
            to_percentage_integer(tb_rate)
        ]

    # ==============================================================================
    # 报表生成系列：常规流程 (01 - 09)
    # ==============================================================================
    def gen_prov_long_merged(self, df_raw, df_dedup, output_file):
        self.log(f"正在生成报表 [01_省份一维]: {os.path.basename(output_file)}")
        groups = BRANCH_GROUPS
        iter_list = [("线上平台", "网站", "网站"), ("线上平台", "小红书", "小红书"), ("线上平台", "公众号", "公众号"), ("线上平台", "抖音", "抖音"), ("线上平台", "视频号", "视频号"), ("线上平台", "其他", "其他"), ("线下活动", "线下活动", "线下活动"), ("考试现场", "考试现场", "考试现场"), ("高校", "高校", "高校"), ("其他", "其他", "其他")]
        
        final_rows = []
        # 初始化全国汇总
        grand_national_stats = {key: {"raw_count":0, "dedup_count":0,"new_added_count":0,"gz_sum":0,"sy_sum":0,"total_tb_sum":0,"reply_count":0} for key in [x[2] for x in iter_list]}
        
        for group_name, branches in groups.items():
            group_accumulator = {key: {"raw_count":0, "dedup_count":0,"new_added_count":0,"gz_sum":0,"sy_sum":0,"total_tb_sum":0,"reply_count":0} for key in [x[2] for x in iter_list]}
            for branch_name in branches:
                branch_raw_df = df_raw[df_raw['分校'] == branch_name]
                branch_dedup_df = df_dedup[df_dedup['分校'] == branch_name]
                
                for p_name, f_keyword, s_name in iter_list:
                    if p_name == "线上平台":
                        current_raw = branch_raw_df[(branch_raw_df['渠道'] == '线上平台') & (branch_raw_df['线上渠道'] == f_keyword)]
                        current_dedup = branch_dedup_df[(branch_dedup_df['渠道'] == '线上平台') & (branch_dedup_df['线上渠道'] == f_keyword)]
                    else:
                        current_raw = branch_raw_df[branch_raw_df['渠道'] == f_keyword]
                        current_dedup = branch_dedup_df[branch_dedup_df['渠道'] == f_keyword]
                    
                    stats = self.calculate_summary_stats(current_raw, current_dedup)
                    
                    for stat_key in stats:
                        group_accumulator[s_name][stat_key] += stats[stat_key]
                        grand_national_stats[s_name][stat_key] += stats[stat_key]
                        
                    final_rows.append([branch_name, p_name, s_name] + self.format_output_row_list(stats, True))
            
            # 写入类别总计行
            for p_name, f_keyword, s_name in iter_list:
                final_rows.append([f"{group_name}类总计", p_name, s_name] + self.format_output_row_list(group_accumulator[s_name], True))
        
        # 写入全国总计行
        for p_name, f_keyword, s_name in iter_list:
            final_rows.append(["全国", p_name, s_name] + self.format_output_row_list(grand_national_stats[s_name], True))
            
        header_names = ["分校", "所属平台", "所属渠道", "新增好友", "本月分校内去重", "净新增", "重复率", "公职标备", "事业标备", "其他标备", "标备总量", "标备率", "回话备注", "回话备注率"]
        pd.DataFrame(final_rows, columns=header_names).to_excel(output_file, index=False)
        self._style_excel(output_file)

    def gen_prov_wide(self, df_raw, df_dedup, output_file, channel_map, strict_online):
        self.log(f"正在生成报表 [省份宽表]: {os.path.basename(output_file)}")
        groups = BRANCH_GROUPS
        final_rows = []
        grand_national_stats = {ct[0]: {"raw_count":0, "dedup_count":0,"new_added_count":0,"gz_sum":0,"sy_sum":0,"total_tb_sum":0,"reply_count":0} for ct in channel_map}
        
        for group_name, branches in groups.items():
            group_accumulator = {ct[0]: {"raw_count":0, "dedup_count":0,"new_added_count":0,"gz_sum":0,"sy_sum":0,"total_tb_sum":0,"reply_count":0} for ct in channel_map}
            for branch_name in branches:
                branch_raw_df = df_raw[df_raw['分校'] == branch_name]
                branch_dedup_df = df_dedup[df_dedup['分校'] == branch_name]
                row_data = [branch_name]
                
                for ct_title, cf_keyword in channel_map:
                    if cf_keyword:
                        if strict_online:
                            current_raw = branch_raw_df[(branch_raw_df['渠道'] == '线上平台') & (branch_raw_df['线上渠道'] == cf_keyword)]
                            current_dedup = branch_dedup_df[(branch_dedup_df['渠道'] == '线上平台') & (branch_dedup_df['线上渠道'] == cf_keyword)]
                        else:
                            current_raw = branch_raw_df[branch_raw_df['渠道'] == cf_keyword]
                            current_dedup = branch_dedup_df[branch_dedup_df['渠道'] == cf_keyword]
                    else:
                        current_raw = branch_raw_df
                        current_dedup = branch_dedup_df
                        
                    stats = self.calculate_summary_stats(current_raw, current_dedup)
                    
                    for stat_key in stats:
                        group_accumulator[ct_title][stat_key] += stats[stat_key]
                        grand_national_stats[ct_title][stat_key] += stats[stat_key]
                        
                    row_data.extend(self.format_output_row_list(stats, True))
                final_rows.append(row_data)
            
            group_summary_row = [f"{group_name}类总计"]
            for ct_title, _ in channel_map:
                group_summary_row.extend(self.format_output_row_list(group_accumulator[ct_title], True))
            final_rows.append(group_summary_row)
            
        national_row = ["全国"]
        for ct_title, _ in channel_map:
            national_row.extend(self.format_output_row_list(grand_national_stats[ct_title], True))
        final_rows.append(national_row)
        
        self._write_wide_excel_structured(output_file, final_rows, channel_map, 2, False)

    def gen_city_long_merged(self, df_raw, df_dedup, df_ret, output_file, do_retention):
        self.log(f"正在生成报表 [04_地市一维]: {os.path.basename(output_file)}")
        iter_list = [("线上平台", "网站", "网站"), ("线上平台", "小红书", "小红书"), ("线上平台", "公众号", "公众号"), ("线上平台", "抖音", "抖音"), ("线上平台", "视频号", "视频号"), ("线上平台", "其他", "其他"), ("线下活动", "线下活动", "线下活动"), ("考试现场", "考试现场", "考试现场"), ("高校", "高校", "高校"), ("其他", "其他", "其他")]
        
        final_rows = []
        grand_national_stats = {key: {"raw_count":0, "dedup_count":0,"new_added_count":0,"gz_sum":0,"sy_sum":0,"total_tb_sum":0,"reply_count":0} for key in [x[2] for x in iter_list]}
        grand_retention_stats = {key: {"gz": 0, "sy": 0, "tb": 0, "ret": 0} for key in [x[2] for x in iter_list]}
        
        all_branches = sorted(df_raw['分校'].dropna().unique())
        for branch_name in all_branches:
            branch_raw_subset = df_raw[df_raw['分校'] == branch_name]
            branch_dedup_subset = df_dedup[df_dedup['分校'] == branch_name]
            branch_retention_subset = df_ret[df_ret['分校'] == branch_name] if do_retention else None
            
            all_cities = sorted(branch_raw_subset['地市'].dropna().unique())
            for city_name in all_cities:
                city_raw_df = branch_raw_subset[branch_raw_subset['地市'] == city_name]
                city_dedup_df = branch_dedup_subset[branch_dedup_subset['地市'] == city_name]
                category_name = city_raw_df['所属类别'].iloc[0] if not city_raw_df.empty else "其它"
                
                for p_name, f_keyword, s_name in iter_list:
                    if p_name == "线上平台":
                        current_raw = city_raw_df[(city_raw_df['渠道'] == '线上平台') & (city_raw_df['线上渠道'] == f_keyword)]
                        current_dedup = city_dedup_df[(city_dedup_df['渠道'] == '线上平台') & (city_dedup_df['线上渠道'] == f_keyword)]
                    else:
                        current_raw = city_raw_df[city_raw_df['渠道'] == f_keyword]
                        current_dedup = city_dedup_df[city_dedup_df['渠道'] == f_keyword]
                    
                    stats = self.calculate_summary_stats(current_raw, current_dedup)
                    for stat_key in stats:
                        grand_national_stats[s_name][stat_key] += stats[stat_key]
                    
                    retention_cols_data = []
                    if do_retention:
                        city_ret_base = branch_retention_subset[branch_retention_subset['地市'] == city_name]
                        if p_name == "线上平台":
                            city_ret_final = city_ret_base[(city_ret_base['渠道'] == '线上平台') & (city_ret_base['线上渠道'] == f_keyword)]
                        else:
                            city_ret_final = city_ret_base[city_ret_base['渠道'] == f_keyword]
                            
                        ret_stats = self.calc_stats_retention_data(city_ret_final)
                        for ret_key in ret_stats:
                            grand_retention_stats[s_name][ret_key] += ret_stats[ret_key]
                            
                        retention_cols_data = [
                            int(ret_stats["ret"]),
                            int(ret_stats["gz"]),
                            int(ret_stats["sy"]),
                            int(ret_stats["tb"] - ret_stats["gz"] - ret_stats["sy"]),
                            int(ret_stats["tb"])
                        ]
                    
                    final_rows.append([branch_name, city_name, category_name, p_name, s_name] + self.format_output_row_list(stats, False) + retention_cols_data)
        
        # 写入全国维度
        for p_name, f_keyword, s_name in iter_list:
            national_row = ["全国", "总计", "-", p_name, s_name] + self.format_output_row_list(grand_national_stats[s_name], False)
            if do_retention:
                rs_national = grand_retention_stats[s_name]
                national_row += [
                    int(rs_national["ret"]),
                    int(rs_national["gz"]),
                    int(rs_national["sy"]),
                    int(rs_national["tb"] - rs_national["gz"] - rs_national["sy"]),
                    int(rs_national["tb"])
                ]
            final_rows.append(national_row)
            
        header_names = ["分校", "地市", "所属类别", "所属平台", "所属渠道", "新增好友", "本月分校内去重", "净新增", "重复率", "净-公职标备", "净-事业标备", "净-其他标备", "净-标备总量", "净-标备率", "回话备注", "回话备注率"]
        if do_retention:
            header_names += ["总-留存量", "总-公职标备", "总-事业标备", "总-其他标备", "总-标备总量"]
            
        pd.DataFrame(final_rows, columns=header_names).to_excel(output_file, index=False)
        self._style_excel(output_file)

    def gen_city_wide(self, df_raw, df_dedup, output_file, channel_map, strict_online):
        self.log(f"正在生成报表 [地市宽表]: {os.path.basename(output_file)}")
        final_rows = []
        grand_national_stats = {ct[0]: {"raw_count":0, "dedup_count":0,"new_added_count":0,"gz_sum":0,"sy_sum":0,"total_tb_sum":0,"reply_count":0} for ct in channel_map}
        
        all_branches = sorted(df_raw['分校'].dropna().unique())
        for branch_name in all_branches:
            branch_raw_subset = df_raw[df_raw['分校'] == branch_name]
            branch_dedup_subset = df_dedup[df_dedup['分校'] == branch_name]
            
            all_cities = sorted(branch_raw_subset['地市'].dropna().unique())
            for city_name in all_cities:
                row_data = [branch_name, city_name]
                city_raw_df = branch_raw_subset[branch_raw_subset['地市'] == city_name]
                city_dedup_df = branch_dedup_subset[branch_dedup_subset['地市'] == city_name]
                
                for ct_title, cf_keyword in channel_map:
                    if cf_keyword:
                        if strict_online:
                            current_raw = city_raw_df[(city_raw_df['渠道'] == '线上平台') & (city_raw_df['线上渠道'] == cf_keyword)]
                            current_dedup = city_dedup_df[(city_dedup_df['渠道'] == '线上平台') & (city_dedup_df['线上渠道'] == cf_keyword)]
                        else:
                            current_raw = city_raw_df[city_raw_df['渠道'] == cf_keyword]
                            current_dedup = city_dedup_df[city_dedup_df['渠道'] == cf_keyword]
                    else:
                        current_raw = city_raw_df
                        current_dedup = city_dedup_df
                        
                    stats = self.calculate_summary_stats(current_raw, current_dedup)
                    for stat_key in stats:
                        grand_national_stats[ct_title][stat_key] += stats[stat_key]
                        
                    row_data.extend(self.format_output_row_list(stats, False))
                final_rows.append(row_data)
                
        national_row = ["全国", "总计"]
        for ct_title, _ in channel_map:
            national_row.extend(self.format_output_row_list(grand_national_stats[ct_title], False))
        final_rows.append(national_row)
        
        self._write_wide_excel_structured(output_file, final_rows, channel_map, 3, True)

    def gen_special_city_report(self, df_raw, df_dedup, output_file, channel_map, strict_online):
        self.log(f"正在生成报表 [独立24地市宽表]: {os.path.basename(output_file)}")
        final_rows = []
        grand_national_stats = {ct[0]: {"raw_count":0, "dedup_count":0,"new_added_count":0,"gz_sum":0,"sy_sum":0,"total_tb_sum":0,"reply_count":0} for ct in channel_map}
        
        for city_name in FIXED_ORDER_CITIES:
            city_raw_df = df_raw[df_raw['地市'] == city_name]
            city_dedup_df = df_dedup[df_dedup['地市'] == city_name]
            branch_name = city_raw_df['分校'].iloc[0] if not city_raw_df.empty else DEFAULT_BRANCH_MAP.get(city_name, "未知分校")
            
            row_data = [branch_name, city_name]
            for ct_title, cf_keyword in channel_map:
                if cf_keyword:
                    if strict_online:
                        current_raw = city_raw_df[(city_raw_df['渠道'] == '线上平台') & (city_raw_df['线上渠道'] == cf_keyword)]
                        current_dedup = city_dedup_df[(city_dedup_df['渠道'] == '线上平台') & (city_dedup_df['线上渠道'] == cf_keyword)]
                    else:
                        current_raw = city_raw_df[city_raw_df['渠道'] == cf_keyword]
                        current_dedup = city_dedup_df[city_dedup_df['渠道'] == cf_keyword]
                else:
                    current_raw = city_raw_df
                    current_dedup = city_dedup_df
                
                stats = self.calculate_summary_stats(current_raw, current_dedup)
                for stat_key in stats:
                    grand_national_stats[ct_title][stat_key] += stats[stat_key]
                    
                row_data.extend(self.format_output_row_list(stats, False))
            final_rows.append(row_data)
            
        national_row = ["全国", "总计"]
        for ct_title, _ in channel_map:
            national_row.extend(self.format_output_row_list(grand_national_stats[ct_title], False))
        final_rows.append(national_row)
        
        self._write_wide_excel_structured(output_file, final_rows, channel_map, 3, True)

    def gen_wide_retention_all_levels(self, df_ret, output_file, channel_map, strict_online, is_prov, is_spec=False):
        self.log(f"正在生成报表 [留存版宽表]: {os.path.basename(output_file)}")
        final_rows = []
        grand_national_stats = {ct[0]: {"gz":0,"sy":0,"tb":0,"ret":0} for ct in channel_map}
        
        if is_spec:
            for city_name in FIXED_ORDER_CITIES:
                city_ret_df = df_ret[df_ret['地市'] == city_name]
                branch_name = city_ret_df['分校'].iloc[0] if not city_ret_df.empty else DEFAULT_BRANCH_MAP.get(city_name, "未知分校")
                row_data = [branch_name, city_name]
                
                for ct_title, cf_keyword in channel_map:
                    if cf_keyword:
                        if strict_online:
                            current_ret = city_ret_df[(city_ret_df['渠道'] == '线上平台') & (city_ret_df['线上渠道'] == cf_keyword)]
                        else:
                            current_ret = city_ret_df[city_ret_df['渠道'] == cf_keyword]
                    else:
                        current_ret = city_ret_df
                        
                    ret_stats = self.calc_stats_retention_data(current_ret)
                    for stat_key in ret_stats:
                        grand_national_stats[ct_title][stat_key] += ret_stats[stat_key]
                        
                    row_data.extend(self.format_retention_row_list(ret_stats))
                final_rows.append(row_data)
                
        elif is_prov:
            groups = BRANCH_GROUPS
            for group_name, branches in groups.items():
                group_stats_accumulator = {ct[0]: {"gz":0,"sy":0,"tb":0,"ret":0} for ct in channel_map}
                for branch_name in branches:
                    branch_ret_df = df_ret[df_ret['分校'] == branch_name]
                    row_data = [branch_name]
                    
                    for ct_title, cf_keyword in channel_map:
                        if cf_keyword:
                            if strict_online:
                                current_ret = branch_ret_df[(branch_ret_df['渠道'] == '线上平台') & (branch_ret_df['线上渠道'] == cf_keyword)]
                            else:
                                current_ret = branch_ret_df[branch_ret_df['渠道'] == cf_keyword]
                        else:
                            current_ret = branch_ret_df
                            
                        ret_stats = self.calc_stats_retention_data(current_ret)
                        for stat_key in ret_stats:
                            group_stats_accumulator[ct_title][stat_key] += ret_stats[stat_key]
                            grand_national_stats[ct_title][stat_key] += ret_stats[stat_key]
                            
                        row_data.extend(self.format_retention_row_list(ret_stats))
                    final_rows.append(row_data)
                    
                group_summary_row = [f"{group_name}类总计"]
                for ct_title, _ in channel_map:
                    group_summary_row.extend(self.format_retention_row_list(group_stats_accumulator[ct_title]))
                final_rows.append(group_summary_row)
                
        else:
            all_branches = sorted(df_ret['分校'].dropna().unique())
            for branch_name in all_branches:
                branch_ret_df = df_ret[df_ret['分校'] == branch_name]
                all_cities = sorted(branch_ret_df['地市'].dropna().unique())
                
                for city_name in all_cities:
                    city_ret_df = branch_ret_df[branch_ret_df['地市'] == city_name]
                    row_data = [branch_name, city_name]
                    
                    for ct_title, cf_keyword in channel_map:
                        if cf_keyword:
                            if strict_online:
                                current_ret = city_ret_df[(city_ret_df['渠道'] == '线上平台') & (city_ret_df['线上渠道'] == cf_keyword)]
                            else:
                                current_ret = city_ret_df[city_ret_df['渠道'] == cf_keyword]
                        else:
                            current_ret = city_ret_df
                            
                        ret_stats = self.calc_stats_retention_data(current_ret)
                        for stat_key in ret_stats:
                            grand_national_stats[ct_title][stat_key] += ret_stats[stat_key]
                            
                        row_data.extend(self.format_retention_row_list(ret_stats))
                    final_rows.append(row_data)
                    
        final_national_row = ["全国", "总计"] if not is_prov else ["全国"]
        for ct_title, _ in channel_map:
            final_national_row.extend(self.format_retention_row_list(grand_national_stats[ct_title]))
        final_rows.append(final_national_row)
        
        self._write_wide_excel_retention_structured(output_file, final_rows, channel_map, 2 if is_prov else 3, not is_prov)

    def gen_date_summary_report_standalone(self, df_dedup, d_str, output_file, is_city):
        if not d_str.strip():
            return
        self.log(f"正在生成报表 [09_日期汇总]: {os.path.basename(output_file)}")
        
        target_dates_list = [d.strip() for d in d_str.replace('，', ',').split(',') if d.strip()]
        filtered_df = df_dedup[df_dedup['日期'].isin(target_dates_list)]
        group_keys_list = ['分校', '地市'] if is_city else ['分校']
        
        count_all = filtered_df.groupby(group_keys_list).size().reset_index(name='total')
        count_web = filtered_df[filtered_df['线上渠道'] == '网站'].groupby(group_keys_list).size().reset_index(name='web')
        summary_result = pd.merge(count_all, count_web, on=group_keys_list, how='left').fillna(0)
        
        sum_total_val = summary_result['total'].sum()
        sum_web_val = summary_result['web'].sum()
        
        national_summary_dict = {col: "" for col in summary_result.columns}
        national_summary_dict['分校'] = "全国"
        if is_city: 
            national_summary_dict['地市'] = "总计"
        national_summary_dict['total'] = sum_total_val
        national_summary_dict['web'] = sum_web_val
        
        summary_result = pd.concat([summary_result, pd.DataFrame([national_summary_dict])], ignore_index=True)
        summary_result.rename(columns={'total': f"汇总({','.join(target_dates_list)})", 'web': "其中:网站"}, inplace=True)
        
        summary_result.to_excel(output_file, index=False)
        self._style_excel(output_file)

    # --- ★★★ 质检核心模块 6.0 最终合体版 ★★★ ---
    def gen_quality_inspection_suite(self, df_clean, df_retention, output_file, remark_exclude, channel_exclude, qc_dates_str):
        self.log(">>> 开始质检综合报表生成业务流...")
        base_filename_str = os.path.splitext(output_file)[0]
        
        # 1. 导出表1 (纯天然底稿，不受日期限制)
        self.log("正在提取表 1：详情明细底稿 (不受日期和业务过滤限制)...")
        df_retention.to_csv(f"{base_filename_str}_表1_详情明细底稿.csv", index=False, encoding='utf-8-sig')
        self.log(f"✅ 表 1 导出成功，共计 {len(df_retention)} 行。")

        # 2. 准备受日期限制的数据池
        target_retention_pool = df_retention.copy()
        target_clean_pool = df_clean.copy()
        
        if qc_dates_str.strip():
            date_list_arr = [d.strip() for d in qc_dates_str.replace('，', ',').split(',') if d.strip()]
            target_retention_pool = target_retention_pool[target_retention_pool['日期'].isin(date_list_arr)]
            target_clean_pool = target_clean_pool[target_clean_pool['日期'].isin(date_list_arr)]
            self.log(f"🔍 已应用质检全局日期过滤: {date_list_arr}")

        # 3. 准备虚假异常池
        df_abnormal_pool = target_retention_pool[
            (target_retention_pool['标备总数'] == 1) & 
            (target_retention_pool['客户回复消息数'] == 0) & 
            (target_retention_pool['好友添加来源'].astype(str).str.contains('扫描渠道二维码', na=False))
        ].copy()
        
        if remark_exclude.strip():
            rk_list = [k.strip() for k in remark_exclude.replace('，', ',').split(',') if k.strip()]
            def filter_remark_fn(val):
                return any(k in str(val) for k in rk_list)
            df_abnormal_pool = df_abnormal_pool[~df_abnormal_pool['备注名'].apply(filter_remark_fn)]
            
        if channel_exclude.strip():
            ck_list = [k.strip() for k in channel_exclude.replace('，', ',').split(',') if k.strip()]
            def filter_channel_fn(val):
                return any(k in str(val) for k in ck_list)
            df_abnormal_pool = df_abnormal_pool[~df_abnormal_pool['添加渠道码'].apply(filter_channel_fn)]

        # ★ 4. 导出表2 虚假明细 CSV
        df_fake_details_final = df_abnormal_pool[df_abnormal_pool['是否同意会话存档'].astype(str).str.strip() == '同意'].copy()
        df_fake_details_final.to_csv(f"{base_filename_str}_表2_虚假备注详情明细.csv", index=False, encoding='utf-8-sig')
        self.log(f"✅ 表 2 (详情明细) 导出成功，共计 {len(df_fake_details_final)} 行。")

        groups = BRANCH_GROUPS

        def get_pct_string(a, b):
            return f"{int(round((a/b)*100))}%" if b != 0 else "0%"

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            self.log("正在渲染质检 Excel 看板 (包含 4 个 Sheet)...")
            
            # --- Sheet 1: 表2汇总 ---
            rows_t2_list = []
            grand_total_t2_count = 0
            
            for group_n, branches in groups.items():
                group_total_count = 0
                for b_name in branches:
                    branch_count = len(df_fake_details_final[df_fake_details_final['分校'] == b_name])
                    group_total_count += branch_count
                    rows_t2_list.append([b_name, branch_count])
                rows_t2_list.append([f"{group_n}类总计", group_total_count])
                grand_total_t2_count += group_total_count
                
            rows_t2_list.append(["全国总计", grand_total_t2_count])
            pd.DataFrame(rows_t2_list, columns=['分校', '虚假备注']).to_excel(writer, sheet_name='2_虚假备注筛选', index=False)

            # --- Sheet 2: 跨分校统计 ---
            df3_base_dedup = target_retention_pool.drop_duplicates(subset=['ExternalUserId', '分校'])
            df3_counts_df = df3_base_dedup.groupby('ExternalUserId').size().reset_index(name='跨分校')
            df3_counts_df.to_csv(f"{base_filename_str}_表3_加几个分校明细底稿.csv", index=False, encoding='utf-8-sig')
            self.log("✅ 表 3 (跨分校明细) 导出成功。")
            
            abnormal_id_list = df3_counts_df[df3_counts_df['跨分校'] >= 10]['ExternalUserId']
            # 回查全量客服添加数
            df4_pool_all_adds = target_retention_pool[target_retention_pool['ExternalUserId'].isin(abnormal_id_list)]
            
            rows_t4_list = []
            grand_total_t4_count = 0
            for group_n, branches in groups.items():
                group_total_count = 0
                for b_name in branches:
                    branch_count = len(df4_pool_all_adds[df4_pool_all_adds['分校'] == b_name])
                    group_total_count += branch_count
                    rows_t4_list.append([b_name, branch_count])
                rows_t4_list.append([f"{group_n}类总计", group_total_count])
                grand_total_t4_count += group_total_count
                
            rows_t4_list.append(["全国总计", grand_total_t4_count])
            pd.DataFrame(rows_t4_list, columns=['分校', '添加客服量']).to_excel(writer, sheet_name='4_跨10个以上分校', index=False)

            # --- Sheet 3: 会话看板 ---
            rows_t5_list = []
            national_acc_t5_dict = {'ret':0, 'has':0, 'no':0, 'db':0}
            
            for group_n, branches in groups.items():
                group_acc_t5_dict = {'ret':0, 'has':0, 'no':0, 'db':0}
                for b_name in branches:
                    branch_group_df = target_retention_pool[target_retention_pool['分校'] == b_name]
                    retention_val = len(branch_group_df)
                    has_reply_val = (branch_group_df['客户回复消息数'] > 0).sum()
                    
                    # 强加同意过滤
                    zero_reply_val = ((branch_group_df['客户回复消息数'] == 0) & (branch_group_df['是否同意会话存档'].astype(str).str.strip() == '同意')).sum()
                    double_zero_val = ((branch_group_df['员工发送消息数'] == 0) & (branch_group_df['客户回复消息数'] == 0) & (branch_group_df['是否同意会话存档'].astype(str).str.strip() == '同意')).sum()
                    
                    group_acc_t5_dict['ret'] += retention_val
                    group_acc_t5_dict['has'] += has_reply_val
                    group_acc_t5_dict['no'] += zero_reply_val
                    group_acc_t5_dict['db'] += double_zero_val
                    
                    national_acc_t5_dict['ret'] += retention_val
                    national_acc_t5_dict['has'] += has_reply_val
                    national_acc_t5_dict['no'] += zero_reply_val
                    national_acc_t5_dict['db'] += double_zero_val
                    
                    rows_t5_list.append([
                        b_name, retention_val, has_reply_val, zero_reply_val, 
                        get_pct_string(zero_reply_val, retention_val), 
                        double_zero_val, 
                        get_pct_string(double_zero_val, retention_val)
                    ])
                    
                rows_t5_list.append([
                    f"{group_n}类总计", group_acc_t5_dict['ret'], group_acc_t5_dict['has'], group_acc_t5_dict['no'], 
                    get_pct_string(group_acc_t5_dict['no'], group_acc_t5_dict['ret']), 
                    group_acc_t5_dict['db'], 
                    get_pct_string(group_acc_t5_dict['db'], group_acc_t5_dict['ret'])
                ])
                
            rows_t5_list.append([
                "全国总计", national_acc_t5_dict['ret'], national_acc_t5_dict['has'], national_acc_t5_dict['no'], 
                get_pct_string(national_acc_t5_dict['no'], national_acc_t5_dict['ret']), 
                national_acc_t5_dict['db'], 
                get_pct_string(national_acc_t5_dict['db'], national_acc_t5_dict['ret'])
            ])
            pd.DataFrame(rows_t5_list, columns=['分校', '留存量', '客户有会话', '客户0会话', '客户0会话率', '双向无会话', '双向无会话率']).to_excel(writer, sheet_name='5_会话情况', index=False)

            # --- Sheet 4: 来源分布 ---
            source_names_list = ["获客助手", "名片分享", "其他", "群聊", "扫描名片二维码", "扫描渠道二维码", "搜索手机号", "微信联系人", "未知"]
            rows_t6_list = []
            national_acc_t6_dict = {s: 0 for s in source_names_list}
            
            for group_n, branches in groups.items():
                group_acc_t6_dict = {s: 0 for s in source_names_list}
                for b_name in branches:
                    branch_group_df = target_clean_pool[target_clean_pool['分校'] == b_name]
                    value_counts_dict = branch_group_df['好友添加来源'].fillna('未知').apply(lambda x: x if x in source_names_list else '其他').value_counts().to_dict()
                    
                    branch_counts_list = [value_counts_dict.get(s, 0) for s in source_names_list]
                    branch_total_sum = sum(branch_counts_list)
                    
                    branch_pcts_list = [get_pct_string(c, branch_total_sum) for c in branch_counts_list]
                    rows_t6_list.append([b_name] + branch_counts_list + [branch_total_sum] + [b_name] + branch_pcts_list + [branch_total_sum])
                    
                    for i, s_name in enumerate(source_names_list):
                        group_acc_t6_dict[s_name] += branch_counts_list[i]
                        national_acc_t6_dict[s_name] += branch_counts_list[i]
                
                group_counts_list = [group_acc_t6_dict[s] for s in source_names_list]
                group_total_sum = sum(group_counts_list)
                group_pcts_list = [get_pct_string(c, group_total_sum) for c in group_counts_list]
                rows_t6_list.append([f"{group_n}类总计"] + group_counts_list + [group_total_sum] + [f"{group_n}类总计"] + group_pcts_list + [group_total_sum])
            
            national_counts_list = [national_acc_t6_dict[s] for s in source_names_list]
            national_total_sum = sum(national_counts_list)
            national_pcts_list = [get_pct_string(c, national_total_sum) for c in national_counts_list]
            rows_t6_list.append(["全国总计"] + national_counts_list + [national_total_sum] + ["全国总计"] + national_pcts_list + [national_total_sum])
            
            headers_t6 = ["分校"] + source_names_list + ["总计"] + ["分校"] + source_names_list + ["总计"]
            pd.DataFrame(rows_t6_list, columns=headers_t6).to_excel(writer, sheet_name='6_好友添加来源', index=False)

        self._style_excel(output_file)
        self.log("🎯 质检汇总分析看板生成完毕！所有任务处理成功。")

    # --- 辅助样式渲染 (保持 4.0 完整合并逻辑) ---
    def _write_wide_excel_structured(self, output_file, rows, channel_map, start_col_idx, has_city_col):
        sub_headers_list = ["新增好友", "本月分校内去重", "净新增", "重复率", "公职标备", "事业标备", "其他标备", "标备总量", "标备率", "回话备注", "回话备注率"]
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            pd.DataFrame(rows).to_excel(writer, index=False, header=False, startrow=2)
            worksheet = writer.sheets['Sheet1']
            worksheet['A1'] = "分校"
            worksheet.merge_cells('A1:A2')
            if has_city_col:
                worksheet['B1'] = "地市"
                worksheet.merge_cells('B1:B2')
                
            current_col_idx = start_col_idx
            for ct_name, _ in channel_map:
                worksheet.cell(row=1, column=current_col_idx).value = ct_name
                for sub_header in sub_headers_list:
                    worksheet.cell(row=2, column=current_col_idx).value = sub_header
                    current_col_idx += 1
                worksheet.merge_cells(start_row=1, start_column=current_col_idx-len(sub_headers_list), end_row=1, end_column=current_col_idx-1)
                
            # ★ 纵向合并单元格逻辑
            if has_city_col:
                max_row_num = worksheet.max_row
                current_branch_name = worksheet['A3'].value
                start_row_num = 3
                for r in range(4, max_row_num + 2):
                    cell_val = worksheet[f'A{r}'].value if r <= max_row_num else None
                    if cell_val != current_branch_name:
                        if start_row_num < r - 1:
                            worksheet.merge_cells(f'A{start_row_num}:A{r-1}')
                        current_branch_name = cell_val
                        start_row_num = r
                        
            self._style_excel_ws(worksheet)

    def _write_wide_excel_retention_structured(self, output_file, rows, channel_map, start_col_idx, has_city_col):
        sub_headers_list = ["留存量", "公职标备", "事业标备", "其他标备", "标备总量", "标备率"]
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            pd.DataFrame(rows).to_excel(writer, index=False, header=False, startrow=2)
            worksheet = writer.sheets['Sheet1']
            worksheet['A1'] = "分校"
            worksheet.merge_cells('A1:A2')
            if has_city_col:
                worksheet['B1'] = "地市"
                worksheet.merge_cells('B1:B2')
                
            current_col_idx = start_col_idx
            for ct_name, _ in channel_map:
                worksheet.cell(row=1, column=current_col_idx).value = ct_name
                for sub_header in sub_headers_list:
                    worksheet.cell(row=2, column=current_col_idx).value = sub_header
                    current_col_idx += 1
                worksheet.merge_cells(start_row=1, start_column=current_col_idx-len(sub_headers_list), end_row=1, end_column=current_col_idx-1)
                
            if has_city_col:
                max_row_num = worksheet.max_row
                current_branch_name = worksheet['A3'].value
                start_row_num = 3
                for r in range(4, max_row_num + 2):
                    cell_val = worksheet[f'A{r}'].value if r <= max_row_num else None
                    if cell_val != current_branch_name:
                        if start_row_num < r - 1:
                            worksheet.merge_cells(f'A{start_row_num}:A{r-1}')
                        current_branch_name = cell_val
                        start_row_num = r
                        
            self._style_excel_ws(worksheet)

    def _style_excel(self, path):
        try:
            wb = load_workbook(path)
            for sheet_name in wb.sheetnames:
                self._style_excel_ws(wb[sheet_name])
            wb.save(path)
        except Exception as e:
            self.log(f"样式渲染跳过 (文件可能被占用): {str(e)}")

    def _style_excel_ws(self, worksheet):
        thin_border = Side(border_style="thin", color="000000")
        full_border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        center_align = Alignment(horizontal="center", vertical="center")
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = full_border
                cell.alignment = center_align
                
        for row in worksheet.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                
        worksheet.column_dimensions['A'].width = 18
        if worksheet['B1'].value == '地市':
            worksheet.column_dimensions['B'].width = 15

# ==============================================================================
# GUI 界面 (分离式业务流设计)
# ==============================================================================
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("数据自动化统计工具 6.1 (正式发布版)")
        self.root.geometry("980x520")
        self.root.configure(bg="#f8f9fa")

        ui_style = ttk.Style()
        ui_style.theme_use('clam')
        ui_style.configure('TNotebook.Tab', font=("Microsoft YaHei", 10, "bold"), padding=[15, 5])
        
        tk.Label(root, text="🚀 数据自动化统计工具 6.1 (终极全量无损版)", font=("Microsoft YaHei", 15, "bold"), bg="#007bff", fg="white").pack(fill="x", pady=0, ipady=10)
        
        main_container = tk.Frame(root, bg="#f8f9fa")
        main_container.pack(fill="both", expand=True, padx=15, pady=10)
        
        left_panel = tk.Frame(main_container, bg="#f8f9fa")
        left_panel.pack(side="left", fill="both", expand=True)
        
        right_panel = tk.Frame(main_container, bg="#f8f9fa", width=350)
        right_panel.pack(side="right", fill="y", padx=(15, 0))
        right_panel.pack_propagate(False)

        self.notebook = ttk.Notebook(left_panel)
        self.notebook.pack(fill="both", expand=True)
        
        self.tab_import = ttk.Frame(self.notebook)
        self.tab_regular = ttk.Frame(self.notebook)
        self.tab_qc = ttk.Frame(self.notebook)
        
        self.notebook.add(self.tab_import, text=" 📂 数据导入 ")
        self.notebook.add(self.tab_regular, text=" 📈 常规报表 ")
        self.notebook.add(self.tab_qc, text=" 📊 质检分析 ")

        # --- 标签页 1：导入 ---
        step1_frame = ttk.LabelFrame(self.tab_import, text=" 步骤 1：导入源文件 ")
        step1_frame.pack(fill="x", pady=20, padx=15, ipady=15)
        
        self.ent_file = ttk.Entry(step1_frame)
        self.ent_file.pack(side="left", fill="x", expand=True, padx=10)
        
        ttk.Button(step1_frame, text="📂 浏览文件", command=self.browse, width=12).pack(side="right", padx=10)

        # --- 标签页 2：常规报表 ---
        step2_frame = ttk.LabelFrame(self.tab_regular, text=" 参数设置 ")
        step2_frame.pack(fill="x", pady=10, padx=10)
        
        param_frame = ttk.Frame(step2_frame)
        param_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(param_frame, text="📅 判断月份:").pack(side="left")
        self.ent_month = ttk.Entry(param_frame, width=15)
        self.ent_month.insert(0, "2026-01")
        self.ent_month.pack(side="left", padx=5)
        
        ttk.Label(param_frame, text="📆 特定日期:").pack(side="left")
        self.ent_date = ttk.Entry(param_frame)
        self.ent_date.pack(side="left", fill="x", expand=True, padx=5)
        
        step3_frame = ttk.LabelFrame(self.tab_regular, text=" 常规任务选择 ")
        step3_frame.pack(fill="x", pady=5, padx=10)
        
        self.var_retention = tk.BooleanVar()
        self.var_prov_long = tk.BooleanVar()
        self.var_prov_wide = tk.BooleanVar()
        self.var_city_long = tk.BooleanVar()
        self.var_city_wide = tk.BooleanVar()
        self.var_special = tk.BooleanVar()
        
        ttk.Checkbutton(step3_frame, text="同时生成“留存版”关联报表", variable=self.var_retention).pack(anchor="w", padx=15)
        
        grid_container = ttk.Frame(step3_frame)
        grid_container.pack(fill="x", padx=15, pady=2)
        ttk.Checkbutton(grid_container, text="01_省份一维表", variable=self.var_prov_long).grid(row=0, column=0, sticky="w", pady=5)
        ttk.Checkbutton(grid_container, text="02_省份宽表", variable=self.var_prov_wide).grid(row=0, column=1, sticky="w", pady=5)
        ttk.Checkbutton(grid_container, text="04_地市一维表", variable=self.var_city_long).grid(row=1, column=0, sticky="w", pady=5)
        ttk.Checkbutton(grid_container, text="05_地市宽表", variable=self.var_city_wide).grid(row=1, column=1, sticky="w", pady=5)
        ttk.Checkbutton(grid_container, text="★ 单独地市专项报表", variable=self.var_special).grid(row=2, column=0, columnspan=2, sticky="w", pady=5)
        
        # 修复点：绑定的对象名必须严格一致
        self.btn_regular = ttk.Button(self.tab_regular, text="📈 执行常规任务", command=lambda: self.run_task('regular'))
        self.btn_regular.pack(side="bottom", fill="x", pady=10, padx=15, ipady=8)

        # --- 标签页 3：质检分析 ---
        step4_frame = ttk.LabelFrame(self.tab_qc, text=" 质检规则配置 ")
        step4_frame.pack(fill="both", expand=True, pady=10, padx=10)
        
        qc_f0 = ttk.Frame(step4_frame)
        qc_f0.pack(fill="x", padx=15, pady=6)
        ttk.Label(qc_f0, text="📆 独立过滤日期:").pack(side="left")
        self.ent_qc_date = ttk.Entry(qc_f0)
        self.ent_qc_date.pack(side="left", fill="x", expand=True, padx=5)
        
        qc_f1 = ttk.Frame(step4_frame)
        qc_f1.pack(fill="x", padx=15, pady=6)
        ttk.Label(qc_f1, text="🚫 备注名排除词:").pack(side="left")
        self.ent_qc_rem = ttk.Entry(qc_f1)
        self.ent_qc_rem.insert(0, "学26,学27,报26,课26,前台,到店,报27,课27")
        self.ent_qc_rem.pack(side="left", fill="x", expand=True, padx=5)
        
        qc_f2 = ttk.Frame(step4_frame)
        qc_f2.pack(fill="x", padx=15, pady=6)
        ttk.Label(qc_f2, text="🚫 渠道码排除词:").pack(side="left")
        self.ent_qc_ch = ttk.Entry(qc_f2)
        self.ent_qc_ch.insert(0, "前台,到店")
        self.ent_qc_ch.pack(side="left", fill="x", expand=True, padx=5)
        
        self.btn_qc_analysis = ttk.Button(self.tab_qc, text="🎯 执行质检分析", command=lambda: self.run_task('qc'))
        self.btn_qc_analysis.pack(side="bottom", fill="x", pady=10, padx=15, ipady=8)

        # --- 右侧：暗黑日志栏 ---
        log_container = ttk.LabelFrame(right_panel, text=" 实时运行日志 ")
        log_container.pack(fill="both", expand=True)
        self.log_text_box = scrolledtext.ScrolledText(log_container, font=("Consolas", 9), bg="#1e1e1e", fg="#d4d4d4", relief="flat")
        self.log_text_box.pack(fill="both", expand=True, padx=2, pady=2)

    def log(self, message):
        """写入日志并强制刷新 UI，防止卡死"""
        self.log_text_box.config(state='normal')
        self.log_text_box.insert(tk.END, f"[{time.strftime('%H:%M:%S')}] {message}\n")
        self.log_text_box.see(tk.END)
        self.log_text_box.config(state='disabled')
        self.root.update_idletasks()

    def browse(self):
        selected_files = filedialog.askopenfilenames(filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")])
        if selected_files:
            self.ent_file.delete(0, tk.END)
            self.ent_file.insert(0, ";".join(selected_files))

    def run_task(self, task_name):
        threading.Thread(target=self.process_logic, args=(task_name,), daemon=True).start()

    def process_logic(self, task_type):
        files_paths = self.ent_file.get().split(";")
        if not files_paths or files_paths[0] == "":
            return messagebox.showerror("错误", "请先在第一页选择数据文件！")
            
        start_timestamp = time.time()
        
        # 按钮锁定状态
        self.btn_regular.config(state="disabled")
        self.btn_qc_analysis.config(state="disabled")
        
        self.log_text_box.config(state='normal')
        self.log_text_box.delete(1.0, tk.END)
        self.log_text_box.config(state='disabled')
        
        try:
            processor_obj = UniversalProcessor(self.log)
            self.log(f"引擎启动，准备读取 {len(files_paths)} 个文件...")
            
            data_frames_list = []
            for f_path in files_paths:
                clean_path = f_path.strip()
                if clean_path.lower().endswith('.csv'):
                    data_frames_list.append(pd.read_csv(clean_path, encoding='utf-8-sig'))
                else:
                    data_frames_list.append(pd.read_excel(clean_path)) # 恢复基础引擎，确保最强兼容性
                    
            full_source_dataframe = pd.concat(data_frames_list, ignore_index=True)
            self.log(f"成功载入数据：{len(full_source_dataframe)} 行。")
            
            clean_df, dedup_prov_df, dedup_city_df, retention_df = processor_obj.process_step1(full_source_dataframe, self.ent_month.get(), "")
            
            base_directory = os.path.dirname(files_paths[0])
            base_filename = os.path.splitext(os.path.basename(files_paths[0]))[0]

            if task_type == 'regular':
                # 常规业务流
                dedup_city_df.to_csv(os.path.join(base_directory, f"{base_filename}_00_源数据备份.csv"), index=False, encoding='utf-8-sig')
                
                map_channel = [("微信好友总量", None), ("线上平台", "线上平台"), ("线下平台", "线下活动"), ("考试现场", "考试现场"), ("高校", "高校"), ("其他", "其他")]
                map_online = [("微信好友总量", None), ("网站", "网站"), ("小红书", "小红书"), ("公众号", "公众号"), ("抖音", "抖音"), ("视频号", "视频号"), ("其他", "其他")]
                
                is_do_retention = self.var_retention.get() and not retention_df.empty
                
                if self.var_prov_long.get():
                    processor_obj.gen_prov_long_merged(clean_df, dedup_prov_df, os.path.join(base_directory, f"{base_filename}_01_省份一维.xlsx"))
                    
                if self.var_prov_wide.get():
                    processor_obj.gen_prov_wide(clean_df, dedup_prov_df, os.path.join(base_directory, f"{base_filename}_02_省份宽表_渠道.xlsx"), map_channel, False)
                    processor_obj.gen_prov_wide(clean_df, dedup_prov_df, os.path.join(base_directory, f"{base_filename}_03_省份宽表_线上.xlsx"), map_online, True)
                    if is_do_retention:
                        processor_obj.gen_wide_retention_all_levels(retention_df, os.path.join(base_directory, f"{base_filename}_02_省宽_渠道_留存.xlsx"), map_channel, False, True)
                        processor_obj.gen_wide_retention_all_levels(retention_df, os.path.join(base_directory, f"{base_filename}_03_省宽_线上_留存.xlsx"), map_online, True, True)
                        
                if self.var_city_long.get():
                    processor_obj.gen_city_long_merged(clean_df, dedup_city_df, retention_df, os.path.join(base_directory, f"{base_filename}_04_地市一维.xlsx"), is_do_retention)
                    
                if self.var_city_wide.get():
                    processor_obj.gen_city_wide(clean_df, dedup_city_df, os.path.join(base_directory, f"{base_filename}_05_地市宽表_渠道.xlsx"), map_channel, False)
                    processor_obj.gen_city_wide(clean_df, dedup_city_df, os.path.join(base_directory, f"{base_filename}_06_地市宽表_线上.xlsx"), map_online, True)
                    if is_do_retention:
                        processor_obj.gen_wide_retention_all_levels(retention_df, os.path.join(base_directory, f"{base_filename}_05_地宽_渠道_留存.xlsx"), map_channel, False, False)
                        processor_obj.gen_wide_retention_all_levels(retention_df, os.path.join(base_directory, f"{base_filename}_06_地宽_线上_留存.xlsx"), map_online, True, False)
                        
                if self.var_special.get():
                    processor_obj.gen_special_city_report(clean_df, dedup_city_df, os.path.join(base_directory, f"{base_filename}_07_单独地市_渠道.xlsx"), map_channel, False)
                    processor_obj.gen_special_city_report(clean_df, dedup_city_df, os.path.join(base_directory, f"{base_filename}_08_单独地市_线上.xlsx"), map_online, True)
                    if is_do_retention:
                        processor_obj.gen_wide_retention_all_levels(retention_df, os.path.join(base_directory, f"{base_filename}_07_单独_渠道_留存.xlsx"), map_channel, False, False, True)
                        processor_obj.gen_wide_retention_all_levels(retention_df, os.path.join(base_directory, f"{base_filename}_08_单独_线上_留存.xlsx"), map_online, True, False, True)
                        
                if self.ent_date.get().strip():
                    processor_obj.gen_date_summary_report_standalone(dedup_city_df, self.ent_date.get(), os.path.join(base_directory, f"{base_filename}_09_日期汇总.xlsx"), True)
            else:
                # 质检业务流
                if retention_df.empty:
                    self.log("❌ 严重错误：未发现任何状态为‘已添加’的留存数据，无法生成质检报表。")
                else:
                    processor_obj.gen_quality_inspection_suite(clean_df, retention_df, os.path.join(base_directory, f"{base_filename}_10_质检汇总看板.xlsx"), self.ent_qc_rem.get(), self.ent_qc_ch.get(), self.ent_qc_date.get())
            
            self.log(f"🎉 全部处理完成！总耗时: {time.time()-start_timestamp:.2f}秒")
            messagebox.showinfo("成功", "任务已全部处理完毕。")
            
        except Exception as e:
            self.log(f"💥 运行异常: {str(e)}")
            messagebox.showerror("运行异常", str(e))
        finally:
            self.btn_regular.config(state="normal")
            self.btn_qc_analysis.config(state="normal")

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
